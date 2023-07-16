###########################################
# 设备点表要求增加连接至哪台交换机
# 根据交换机点表，提取其楼层、编号、IP
# 写入摄像机点表中
###########################################


import openpyxl as xl
from openpyxl.styles import Alignment

switchs = {}

book = xl.load_workbook('./excel/点表（分楼层）（室外相机改室内）.xlsx')
floors = book.sheetnames

book2 = xl.load_workbook('安防交换机实施表格0511.xlsx')
sheet2 = book2['Sheet2']

max_r = sheet2.max_row + 1
# 先提取交换机信息加入字典switchs
for row in range(1, max_r):
    switch_floor = sheet2[f'A{row}'].value
    if switch_floor is not None:
        switch_serial = sheet2[f'C{row}'].value
        switch_ip = sheet2[f'E{row}'].value

        s = switch_floor + '\n' + switch_serial + '\n' + switch_ip
        switchs[s] = switch_floor

# 将信息写入点表中
for floor in floors:
    sheet = book[floor]
    first_row = 2
    row = 2
    for switch in switchs:
        if switchs[switch] == floor:
            sheet[f'A{first_row}'].value = switch
            sheet[f'A{first_row}'].alignment = Alignment(wrapText=True)
            sheet.merge_cells(f'A{first_row}:A{first_row + 23}')
            first_row += 24
            for n in range(1, 25):
                sheet[f'B{row}'].value = n
                row += 1

book.save('./excel/点表_test.xlsx')
