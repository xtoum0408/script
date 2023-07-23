import openpyxl as xl
from openpyxl.styles import Font, colors, Alignment

book = xl.load_workbook('D:/excel/camera.xlsx')
floors = book.sheetnames
font = Font(color=colors.BLUE, underline='single')
camera_name = ['人脸8249E-ZRL', '半球8239E-Z', '室外人脸枪机9828P',
               '球D20T', '室外枪HF952', '高低空', '电梯', '4k',
               '室外球9430UA', '热成像8421', '室外热成像']
camera_type_serial = ['DH-IPC-HDBW8249E-ZRL', 'DH-IPC-HDBW8239E-Z', 'DH-SH-HF-9828P-I',
                      'DH-SH-SD9D20T-HNI', 'DH-IPC-HF952', 'DH-IPC-PFW8808-A180', 'DH-SH-HDBW9121P',
                      'DH-SH-HF9868P-I', 'DH-SD-6A9430UA-HNI', 'DH-TPC-SD8421', 'DH-TPC-BF7121']


def camera_type_confirm(camera):
    global t
    for cn in camera_name:
        if camera == cn:
            t = camera_type_serial[camera_name.index(cn)]
            # print(t)
            break
    return t


for floor in floors:
    sheet = book[floor]
    max_r = sheet.max_row + 1
    sheet['D1'].value = '结论'
    for row in range(2, max_r):
        if sheet[f'A{row}'].value is None:break
        serial = sheet[f'B{row}'].value
        sheet[f'D{row}'].value = '对焦正常，画面清晰'
        sheet[f'E{row}'].value = f'=HYPERLINK("./pic/{floor}/{serial}.png","查看结果")'
        sheet[f'E{row}'].font = font
        sheet[f'C{row}'].value = camera_type_confirm(sheet[f'C{row}'].value)

book.save('D:/excel/camera_with_link2.xlsx')
print('-----------完成！！！！-----------')
