#######################################
# 按照模板导入相机的信息
#######################################

# 导入库
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from openpyxl import Workbook
import openpyxl as xl
import os

# 定义变量
user = 'admin'
key = 'xinkaifa9'
evs_name = ''
evs_IP = ''
channel_number = 30
channel_name = ''
camera = ''
code = 10000

title = {
            '所属组织编码': '#001',
            '用户': user,
            '密码': key,
            '厂商': 'Dahua',
            '设备序列号':'',
            '描述':'',
            '所属组织':'新开发银行',
            '设备名称': evs_name,
            '设备类型': 'EVS',
            '所属视频服务器': '中心服务器',
            'IP地址': evs_IP,
            '端口号': 37777,
            '注册服务':'',
            '代理端口': 61025,
            '主动注册':'',
            '添加方式':'IP/IP段',
            '存储回放服务':'',
            '码流':'主码流',
            '转码':'否',
            '零通道编码':'否',
            '通道类别':'视频通道',
            '通道数':channel_number,
            '通道名称':channel_name,
            '是否启用':'是',
            '摄像头':camera,
            '唯一标识码':code,
            '经度':'',
            '纬度':'',
            '防线':'',
            '故障原因':'',
            '放射类型':'',
            '照视角度':'',
            '照视范围':'',
            '照视距离(M)':'',
            '是否支持可视域':'',
            '是否支持全景云台':'',
            '报警类型':'',
            '报警等级':''
        }

file_name = '各楼层摄像机编号和IP_存储分区用.xlsx'
camera_data = {}

no = 1  # 循环计数
title_keys = []
camera_temp = {}
camera_ip_for_key = []

# 读取全部相机数据装进一个字典里，相机编号为key，ip和类型做list
book = xl.load_workbook(file_name)
floors = book.sheetnames
for floor in floors:
    sheet = book[floor]
    max_r = sheet.max_row + 1

    for row in range(2, max_r):
        if sheet[f'A{row}'].value is None:
            break
        ip = sheet[f'A{row}'].value
        camera_ip_for_key.append(ip)
        serial = sheet[f'B{row}'].value
        c = sheet[f'C{row}'].value
        if c == '球D20T' or '室外球9430UA' or '热成像8421':
            camera_type = 'Speed Dome'
        elif c == '室外枪HF952' or '室外人脸枪机9828P':
            camera_type = 'Fixed Camera'
        else:
            camera_type = 'Dome Camera'
        camera_data[ip] = [serial, camera_type]

# print(len(camera_data))

# 新建一个表格
target_book = Workbook()
target_sheet = target_book.active
target_sheet.title = 'devices'
target_sheet = target_book['devices']

# 标题行
for key in title.keys():
    title_keys.append(key)

for lie in range(0,38):
    target_sheet.cell(row=1,column=lie+1).value = title_keys[lie]
    # print(target_sheet.cell(row=1,column=lie+1).value)
# 循环通道数
first_camera_no = 0
last_camera_no = channel_number - 1
camera_no = 0
count = 1
row = 2
while camera_no <= 1568:
    # 按30个一组将数组存入临时dict
    for number in range(0, channel_number):
        if camera_no <= 1558:
            key_ip = camera_ip_for_key[camera_no]
            camera_temp[key_ip] = camera_data[key_ip]
        else:
            break
        camera_no = camera_no + 1

    # 将数据填入表格
    title['设备名称'] = f'evs{count}'
    title['IP地址'] = f'10.11.11.{count}'
    count = count + 1

    for camera_key in camera_temp:
        channel_name = camera_key
        title['通道名称'] = camera_temp[camera_key][0]
        title['摄像头'] = camera_temp[camera_key][1]
        code = code + 1
        title['唯一标识码'] = code
        for column in range(1, 39):
            sheet_title = target_sheet.cell(row=1,column=column).value
            target_sheet.cell(row=row, column=column).value = title[sheet_title]
        row = row + 1
    camera_temp.clear()

# 最后两行加报警行
    target_sheet[f'U{row}'].value = '报警输入通道'
    target_sheet[f'V{row}'].value = '0'
    row = row + 1
    target_sheet[f'U{row}'].value = '报警输出通道'
    target_sheet[f'V{row}'].value = '0'
    row = row + 1
# 保存表格
target_book.save('device.xlsx')
