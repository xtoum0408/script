#############################################
# 表格内通道名称需修改成安装位置以匹配核心设备工作要求
# ###########################################

import openpyxl as xl

book = xl.load_workbook('./excel/管理平台配置表（旧表）.xlsx')
sheet = book['device']
max_r = sheet.max_row + 1

book2 = xl.load_workbook('./excel/相机点表（修改通道名称用）.xlsx')
sheet2 = book2['弱电']
max_r2 = sheet2.max_row + 1

serials = []
location = []
count = 0

# 把点表的信息放进list，避免多次读取表格
for row in range(2, max_r2):
    value = sheet2[f'G{row}'].value
    if value is not None:
        serials.append(value)
        location.append(sheet2[f'D{row}'].value)

# 修改配置表中的内容
for row in range(2, max_r):
    value = sheet[f'W{row}'].value
    if value is not None:
        if value in serials:
            index = serials.index(value)
            sheet[f'W{row}'].value = location[index]
            print(location[index])
            count += 1

book.save('./excel/管理平台配置表（含安装位置）.xlsx')
print("完成！！！")

