# 陈双玥 2020年8月17日
# 注意修改楼层
# 注意摄像机编号均为两位数，如“01”“02”
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from openpyxl import Workbook
import openpyxl as xl
import os

key = "xinkaifa9"
user = "admin"
mac_add = str()
cells = []
wangguan = ['10', '11', '0', '254']

# floor = str(input('请输入楼层：'))  # 注意修改F
floor = 'B3'
# serial = str(input('请输入摄像机序号：'))

# 读取excel
# book = xl.load_workbook("D:\工作\安装集团\摄像头编号\(总)各楼层摄像机编号和IP.xlsx")
book = xl.load_workbook("(总)各楼层摄像机编号和IP.xlsx")

sheet = book[floor]
max_r = sheet.max_row  # 最大行
max_c = sheet.max_column  # 最大列
r = int(max_r - 1)
# print(max_r, max_c)
# print(book.sheetnames)


serial = str(input('请输入摄像机序号：'))
for i in range(0, r):
        full_serial = sheet[f'B{i + 2}'].value  # 获得完整编号
        cells.append(full_serial)
        serial_list = cells[i].split('-')
        if serial_list[3] == serial:  # 最后一位是否与目标摄像机编号匹配
        # if serial_list[2] == serial:  # 外总体
            target_row = i + 2
            break

    # print(sheet[f'B{target_row}'].value)
ip = str(sheet[f'A{target_row}'].value)  # 获得对应ip地址
ip_list = ip.split('.')
print(ip)
print(full_serial)


address = 'http://' + ip
chromedriver = "chromedriver.exe"  # 这里写本地的chromedriver 的所在路径
os.environ["webdriver.Chrome.driver"] = chromedriver  # 调用chrome浏览器
driver = webdriver.Chrome(chromedriver)

driver.get(address)
# driver.get("http://192.168.1.108")  # 该处为具体网址
driver.refresh()  # 刷新页面


driver.maximize_window() # 浏览器最大化

# 登录
driver.find_element_by_id("login_user").send_keys(user)
driver.find_element_by_id("login_psw").send_keys(key)
sleep(0.5)
driver.find_element_by_xpath("//div[@id='login']/div[1]/div/div[2]/form/div[6]/a[1]").click()
sleep(2)
# waitKey = input("输入任意内容继续：")

'''
# 设置
driver.find_element_by_xpath("//div[@id='main']/ul/li[7]/span").click()
# driver.find_element_by_xpath("//div[@id='main']/ul/li[6]/span").click()   # 人脸识别枪机
sleep(1)

# 视频
driver.find_element_by_xpath("//*[@id='set-menu']/li[1]/ul/li[2]/span").click()
sleep(2)

# 视频叠加
driver.find_element_by_xpath("//*[@id='page_encodeConfig']").is_enabled()
driver.find_element_by_xpath("//*[@id='page_encodeConfig']/ul").is_enabled()
driver.find_element_by_xpath("//*[@id='page_encodeConfig']/ul/li[3]").click()
sleep(1)

# 视频通道
driver.find_element_by_xpath("//*[@id='page_encodeConfig']/div/div[3]/div[2]").is_enabled()
driver.find_element_by_xpath("//*[@id='osd_content']/ul").is_enabled()
driver.find_element_by_xpath("//*[@id='osd_content']/ul/li[2]/span").click()  # 通道标题
driver.find_element_by_xpath("//*[@id='osd_content']/ul/li[2]/span").click()
driver.find_element_by_xpath("//*[@id='osd_channel_title']/ul/li/input").clear()
driver.find_element_by_xpath("//*[@id='osd_channel_title']/ul/li/input").send_keys(full_serial)
driver.find_element_by_id("video_OSD_confirm").click()
sleep(1)


driver.find_element_by_xpath("//*[@id='main']/ul/li[1]/span").click()
'''

