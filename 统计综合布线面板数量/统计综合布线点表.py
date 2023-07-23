# 统计各类型的面板数量
import openpyxl as xl

book = xl.load_workbook('./excel/综合布线点表.xlsx')
floors = book.sheetnames
print('楼层-双口网络-双口电话-单口网络-单口电话-海事-保密')

for floor in floors:

    mokuai = 0
    swmianban = 0
    sdmianban = 0
    dwmianban = 0
    ddmianban = 0
    haishi = 0
    baomi = 0

    sheet = book[floor]
    max_r = sheet.max_row + 1

    for row in range(4, max_r):
        asset = sheet[f'F{row}'].value
        if asset is not None:
            if '双口' in asset and '网络' in asset:
                swmianban += 1
                # mokuai += 2
            elif '双口' in asset and '电话' in asset:
                sdmianban += 1
                # mokuai += 2
            elif '单口' in asset and '网络' in asset:
                dwmianban += 1
                # mokuai += 2
            elif '单口' in asset and '电话' in asset:
                ddmianban += 1
                # mokuai += 1
            elif '海事' in asset:
                haishi += 1
            elif '保密' in asset:
                baomi += 1
            else:
                pass


    print(floor, '-', swmianban, '-', sdmianban, '-', dwmianban, '-', ddmianban,'-', haishi,'-', baomi)
