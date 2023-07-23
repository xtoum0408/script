##############################################################
# 原表中J-N列的单元格的合并情况较为混乱，根据模板进行重新合并和整理
############################################################


import openpyxl as xl
from openpyxl.styles import Alignment
book = xl.load_workbook('test.xlsx')
sheet = book['报警']
max_r = sheet.max_row + 1
A = 1
D = 1
num = 1
merged = []
n = 1

m = sheet.merged_cells
for temp in m:
    if 'K' in str(temp):
        t = str(temp).replace('K', 'J')
        merged.append(t)
        # print(str(temp))

for r in merged:
    start = r.split(':')[0]
    end = r.split(':')[1]
    sheet.merge_cells(f'{start}:{end}')

for row in range(4, 422):
    value = sheet[f'K{row}'].value
    if value is not None:
        sheet[f'J{row}'].value = 'IA-A' + str(num).zfill(2)
        num += 1

for row in range(4, 422):
    value = sheet[f'J{row}'].value
    if value is not None:
        n = 1
        first = row
        sheet[f'I{row}'].value = value + '/D' + str(n)
        n += 1
    else:
        sheet[f'I{row}'].value = sheet[f'J{first}'].value + '/D' + str(n)
        n += 1


book.save('test2.xlsx')
