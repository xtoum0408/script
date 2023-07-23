# 陈双玥 2020年8月17日
# 注意修改楼层
# 注意摄像机编号均为两位数，如“01”“02”
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from openpyxl import Workbook
import openpyxl as xl
import os

key = "xinkaifa9"
user = "admin"
mac_add = str()
cells = []
wangguan = ['10', '11', '0', '254']

# floor = str(input('请输入楼层：'))  # 注意修改F
floor = '南B4'
serial = str(input('请输入摄像机序号：'))

# 读取excel
# book = xl.load_workbook("D:\工作\安装集团\摄像头编号\(总)各楼层摄像机编号和IP.xlsx")
book = xl.load_workbook("(总)各楼层摄像机编号和IP.xlsx")

sheet = book[floor]
max_r = sheet.max_row  # 最大行
max_c = sheet.max_column  # 最大列
r = int(max_r - 1)
# print(max_r, max_c)
# print(book.sheetnames)

for i in range(0, r):
    full_serial = sheet[f'B{i + 2}'].value  # 获得完整编号
    cells.append(full_serial)
    serial_list = cells[i].split('-')
    if serial_list[3] == serial:  # 最后一位是否与目标摄像机编号匹配
    # if serial_list[2] == serial:  # 外总体
        target_row = i + 2
        break

# print(sheet[f'B{target_row}'].value)
ip = str(sheet[f'A{target_row}'].value)  # 获得对应ip地址
ip_list = ip.split('.')
print(ip_list)
print(full_serial)

# 初始化流程

# address = 'http://' + input('请输入ip地址：')
# chromedriver = "chromedriver.exe"  # 这里写本地的chromedriver 的所在路径
# os.environ["webdriver.Chrome.driver"] = chromedriver  # 调用chrome浏览器
driver = webdriver.Firefox()

# driver.get(address)
driver.get("http://192.168.1.108")  # 该处为具体网址
driver.refresh()  # 刷新页面


# driver.maximize_window() #浏览器最大化

# 初始化设置密码
def init():
    driver.find_element_by_name("newpwd").send_keys(key)
    driver.find_element_by_name("newpwdcfm").send_keys(key)
    driver.find_element_by_id("devInit_phone_enable").click()
    driver.find_element_by_xpath("//*[@id='device_init']/div[3]/div/a").click()
    sleep(4)

    driver.find_element_by_name("licence_check").click()
    driver.find_element_by_xpath("//div[@id='login_permission1']/div[3]/div/a").click()
    sleep(1)

    driver.find_element_by_xpath("//div[@id='login_permission2']/div[3]/div/a").click()
    sleep(1)

    driver.find_element_by_xpath("//div[@id='login_permission3']/div[3]/div/a").click()
    sleep(3)
    return


init()

# 登录
driver.find_element_by_id("login_user").send_keys(user)
driver.find_element_by_id("login_psw").send_keys(key)
sleep(0.5)
driver.find_element_by_xpath("//div[@id='login']/div[1]/div/div[2]/form/div[6]/a[1]").click()
sleep(2)
# waitKey = input("输入任意内容继续：")

# 设置
driver.find_element_by_xpath("//div[@id='main']/ul/li[7]/span").click()
# driver.find_element_by_xpath("//div[@id='main']/ul/li[6]/span").click()   # 人脸识别枪机
sleep(1)

# 视频
driver.find_element_by_xpath("//*[@id='set-menu']/li[1]/ul/li[2]/span").click()
sleep(2)
# 视频叠加
driver.find_element_by_xpath("//*[@id='page_encodeConfig']").is_enabled()
driver.find_element_by_xpath("//*[@id='page_encodeConfig']/ul").is_enabled()
driver.find_element_by_xpath("//*[@id='page_encodeConfig']/ul/li[3]").click()
sleep(1)

# 视频通道
driver.find_element_by_xpath("//*[@id='page_encodeConfig']/div/div[3]/div[2]").is_enabled()
driver.find_element_by_xpath("//*[@id='osd_content']/ul").is_enabled()
driver.find_element_by_xpath("//*[@id='osd_content']/ul/li[2]/span").click()  # 通道标题
driver.find_element_by_xpath("//*[@id='osd_content']/ul/li[2]/span").click()
driver.find_element_by_xpath("//*[@id='osd_channel_title']/ul/li/input").clear()
driver.find_element_by_xpath("//*[@id='osd_channel_title']/ul/li/input").send_keys(full_serial)
driver.find_element_by_id("video_OSD_confirm").click()
sleep(1)

# 网络设置
driver.find_element_by_xpath("//*[@id='set-menu']").is_enabled()
driver.find_element_by_xpath("//ul[@id='set-menu']/li[2]/a/span").click()
driver.find_element_by_xpath("//ul[@id='set-menu']/li[2]/a/span").is_enabled()

# tcp/ip
driver.find_element_by_xpath("//ul[@id='set-menu']/li[2]/ul/li[1]/span").click()
driver.find_element_by_xpath("//*[@id='page_networkConfig']").is_enabled()
sleep(3)

# 获得mac地址
driver.find_element_by_xpath("//*[@id='NN_macadd']").is_enabled()
for n in range(1, 7):
    driver.find_element_by_xpath(f"//div[@id='NN_macadd']/input[{n}]").is_enabled()
    mac = driver.find_element_by_xpath(f"//div[@id='NN_macadd']/input[{n}]").get_attribute('value')
    mac_add = mac_add + mac + '.'

print(mac_add[:-1])
sheet[f'E{target_row}'] = mac_add[:-1]

# book.save("D:\工作\安装集团\摄像头编号\mac.xlsx")
# sleep(1)

# 修改ip
for i in range(0, 4):
    ip00 = driver.find_element_by_xpath(f"//*[@id='NN_IPV4_IP']/input[{i + 1}]")
    ActionChains(driver).double_click(ip00).perform()
    ip00.send_keys(ip_list[i])

# 修改子网掩码
ma = driver.find_element_by_xpath("//*[@id='NN_IPV4_SM']/input[3]")
ActionChains(driver).double_click(ma).perform()
ma.send_keys(248)
# ma.send_keys(192) # 保险库档案库

# 修改网关
for i in range(0, 4):
    wang = driver.find_element_by_xpath(f"//*[@id='NN_IPV4_DG']/input[{i + 1}]")
    ActionChains(driver).double_click(wang).perform()
    wang.send_keys(wangguan[i])

sleep(3)
driver.find_element_by_xpath("//*[@id='page_networkConfig']/div/div/div[15]/a[3]").click()
sleep(5)
driver.close()

