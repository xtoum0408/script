import openpyxl as xl

dianbiao_first_line = 2
no = 1
book = xl.load_workbook('(总)各楼层摄像机编号和IP.xlsx')
dianbiao = xl.load_workbook('相机点表.xlsx')

camera_name = ['人脸8249E-ZRL', '半球8239E-Z', '室外人脸枪机9828P',
               '球D20T', '室外枪HF952', '高低空', '电梯', '4k',
               '室外球9430UA', '热成像8421', '室外热成像']
camera_type_serial = ['DH-IPC-HDBW8249E-ZRL', 'DH-IPC-HDBW8239E-Z', 'DH-SH-HF-9828P-I',
                      'DH-SH-SD9D20T-HNI', 'DH-IPC-HF952', 'DH-IPC-PFW8808-A180', 'DH-SH-HDBW9121P',
                      'DH-SH-HF9868P-I', 'DH-SD-6A9430UA-HNI', 'DH-TPC-SD8421', 'DH-TPC-BF7121']

camera_name_CN = ['高清1080P半球（人脸抓拍）', '高清1080P半球', '室外高清1080P枪机（人脸抓拍）',
                  '室内1080P高速智能球', '室外高清1080P枪机', '室外一体化全景相机', '高清1080P电梯专用相机', '4K摄像机',
                  '室外1080P高速智能球', '室外一体化双目热成像摄像机', '室外一体化双目热成像摄像机']

floors = book.sheetnames
sheet_dianbiao = dianbiao['弱电']

for floor in floors:
    sheet = book[floor]
    max_r = sheet.max_row + 1
    for row in range(2, max_r):
        if sheet[f'A{row}'].value == None:
            break
        # 编号
        sheet_dianbiao[f'A{dianbiao_first_line}'].value = no
        no += 1
        # 楼层
        sheet_dianbiao[f'C{dianbiao_first_line}'].value = floor
        # 安装位置
        if 'B' in floor:
            pos = '停车场'
        else:
            pos = '走道'

        sheet_dianbiao[f'D{dianbiao_first_line}'].value = pos

        # 系统
        sheet_dianbiao[f'E{dianbiao_first_line}'].value = '弱电'
        # 设备名称描述
        serial = sheet[f'C{row}'].value
        index = camera_name.index(serial)
        sheet_dianbiao[f'F{dianbiao_first_line}'].value = camera_name_CN[index]
        # sheet_dianbiao[f'F{dianbiao_first_line}'].value = camera_name_CN[camera_name.index(sheet[f'C{row}'].value)]
        # 设备编号
        sheet_dianbiao[f'G{dianbiao_first_line}'].value = sheet[f'B{row}'].value
        # 数量
        sheet_dianbiao[f'H{dianbiao_first_line}'].value = 1
        # 资产详细信息
        sheet_dianbiao[f'K{dianbiao_first_line}'].value = sheet[f'A{row}'].value
        # 资产制造商
        sheet_dianbiao[f'M{dianbiao_first_line}'].value = '华三'
        # 资产型号
        sheet_dianbiao[f'N{dianbiao_first_line}'].value = camera_type_serial[camera_name.index(sheet[f'C{row}'].value)]

        dianbiao_first_line += 1

print('--------完成！！--------')
dianbiao.save('dianbiao.xlsx')