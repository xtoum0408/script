##############################################################
# 约有两百多台相机的人脸功能未开启，通过脚本批量进入网页控制端开启
###############################################################

# 导入库
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from openpyxl import Workbook
import openpyxl as xl
import os

# 定义变量
user = 'admin'
key = 'xinkaifa9'
file_name = '(总)各楼层摄像机编号和IP.xlsx'
camera_type = '人脸8249E-ZRL'
face_id_list = []  # 人脸识别的相机ip地址list
i = 0

# 遍历表格，确定人脸识别相机的ip地址，并赋值给一个list
book = xl.load_workbook(file_name)
floors = book.sheetnames
# print(floors)
for floor in floors[1:]:
    sheet = book[floor]
    max_r = sheet.max_row + 1

    for row in range(2, max_r):
        if sheet[f'C{row}'].value == camera_type:
            face_id_list.append(sheet[f'A{row}'].value)
            i = i + 1

# print(face_list)
# print(i)

# 循环list进行人脸设置
for ip in face_id_list:
    print ('当前IP地址为：', ip)
    driver = webdriver.Firefox()
    address = f'http://{ip}'
    driver.get(address)
    driver.refresh()
    driver.maximize_window()

    # 登录
    driver.find_element_by_id("login_user").send_keys(user)
    driver.find_element_by_id("login_psw").send_keys(key)
    sleep(0.5)
    driver.find_element_by_xpath("//div[@id='login']/div[1]/div/div[2]/form/div[6]/a[1]").click()
    sleep(2)

    # 设置
    driver.find_element_by_xpath("//div[@id='main']/ul/li[7]/span").click()
    # driver.find_element_by_xpath("//div[@id='main']/ul/li[6]/span").click()   # 人脸识别枪机
    sleep(1)

    # 事件管理
    driver.find_element_by_xpath("//*[@id='set-menu']").is_enabled()
    driver.find_element_by_xpath("//ul[@id='set-menu']/li[3]/a/span").click()
    driver.find_element_by_xpath("//ul[@id='set-menu']/li[3]/a/span").is_enabled()

# 智能方案
    driver.find_element_by_xpath("//*[@id='set-menu']/li[3]/ul/li[3]/span").click()
    driver.find_element_by_id('page_iExclusionConfig').is_enabled()
    driver.find_element_by_xpath("//*[@id='page_iExclusionConfig']/div/div[3]/ul/li[1]/div[1]").click()
    driver.find_element_by_xpath("//*[@id='page_iExclusionConfig']/div/div[3]/div[1]/a[3]").click()
    sleep(1)

# 人脸检测（启用→人脸增强→启动人脸曝光→检测区绘制→确定）
    driver.find_element_by_xpath ("//*[@id='set-menu']/li[3]/ul/li[5]/span").click ()
    driver.find_element_by_id('page_ipcFaceNewConfig').is_enabled()
    sleep(4)
    driver.find_element_by_id('i_f_enable').click()
    sleep(0.5)
    driver.find_element_by_id('i_f_enhance').click()
    sleep(0.5)
    driver.find_element_by_id('i_f_expo_enable').click()
    sleep(0.5)
    driver.find_element_by_xpath("//*[@id='page_ipcFaceNewConfig']/div/div/div[1]/div[3]/a[1]").click()
    sleep(0.5)
    driver.find_element_by_xpath("//*[@id='page_ipcFaceNewConfig']/div/div/div[2]/div/div[25]/a[3]").click()
    sleep(5)
    driver.close()
    print('-----------DONE-----------')


