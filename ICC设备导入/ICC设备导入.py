######################################
# 需要往大华的核心设备内添加连接的摄像头的信息
# 已有的信息是EVS及EVS内添加的设备
# 将这些信息导入核心设备提供的模板
######################################

import openpyxl as xl
import random
IP = ''
name = ''
serial = ''
channel = ''
camera_type = ''
infor = [
    '大华协议', '大华厂商', 'IP方式添加', '编码器', IP, '37777',
    '', '', 'admin', 'xinkaifa9', '根节点', '10.11.11.80', name, 'EVS',
    serial, '', '', '视频通道', '辅码流2', channel, '视频通道', camera_type,
    '', '', '电动聚焦', '', '0'
]

channels = []
camera_types = []
nums = []

book = xl.load_workbook('./excel/device0621.xlsx')
sheet = book['device']
mar_r = sheet.max_row + 1

for row in range(2, mar_r):
    v1 = sheet[f'W{row}'].value # 通道名
    v2 = sheet[f'Y{row}'].value # 摄像头类型
    v3 = sheet[f'V{row}'].value # 通道数
    if v1 is not None:
        channels.append(v1)
        camera_types.append(v2)
        if v3 is not None:
            nums.append(v3)

book = xl.load_workbook('./excel/admin_视频设备_模板.xlsx')
sheet = book['admin']
index = 0
r = 2
# 先填写除了evs的ip以外的内容
for channel in channels:
    camera_type = camera_types[index]
    list = infor.copy()

    list[19] = channel
    list[21] = camera_type
    c = 1
    for v in list:
        sheet.cell(row=r,column=c).value = v
        c += 1
    r += 1
    index += 1

# 填写evs和设备唯一标识码
first_row = 2
n = 1
for row in nums:
    last_row = first_row + row + 1
    N = str(n)
    serial = '6F0' + N.rjust(4, '0') + 'PAJ0' + N.rjust(4, '0')
    print(serial, first_row, last_row)
    for r in range(first_row, last_row):
        sheet[f'E{r}'].value = '10.11.11.' + N
        sheet[f'M{r}'].value = '10.11.11.' + N
        sheet[f'O{r}'].value = serial
    n += 1
    first_row = last_row - 1


book.save('admin_视频设备.xlsx')
# print(nums)

